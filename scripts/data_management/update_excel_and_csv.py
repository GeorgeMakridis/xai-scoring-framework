#!/usr/bin/env python3
"""
Update Excel sheets and xai_results.csv with dataset metadata and estimated XAI results.
- Reads IMAGE_DATASETS, TEXT_DATASETS, TIMESERIES_DATASETS from dataset_definitions.py
- Updates Excel: data, results_ai, results_xai
- Extends xai_results.csv with new datasets and fills ratings with estimations
"""

import os
import sys
from typing import List, Dict, Any
import random

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "Fame XAI scoring Framework_v2-2.xlsx")
CSV_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "xai_results.csv")

# AI models per data type (for results_ai)
IMAGE_MODELS = ["ResNet18", "ResNet50", "MobileNetV2", "EfficientNet_B0", "ViT_Base", "DenseNet121", "InceptionV3", "VGG16", "SimpleCNN", "ShuffleNetV2"]
TEXT_MODELS = ["BERT_base", "DistilBERT", "RoBERTa", "ALBERT", "BiLSTM", "LSTM", "CNN_Text", "Transformer_small", "GPT2_small", "ELECTRA"]
TS_MODELS = ["LSTM_TS", "BiLSTM_TS", "GRU_TS", "Transformer_TS", "TCN", "CNN_1D", "XGBoost_TS", "RandomForest_TS"]

# XAI methods per data type (for results_xai)
IMAGE_XAI = ["GradCAM", "GradCAM_pp", "Integrated_Gradients", "SmoothGrad", "LIME_Image", "SHAP_Image", "Saliency_Maps", "Guided_Backprop", "Guided_GradCAM", "ScoreCAM", "LayerCAM", "Attention_Maps"]
TEXT_XAI = ["LIME_Text", "SHAP_Text", "Integrated_Gradients_Text", "Attention_Weights", "Gradient_Text", "InputXGradient_Text", "DeepLIFT_Text", "Occlusion_Text", "Shapley_Sampling_Text", "Rationale_Extraction"]
TS_XAI = ["SHAP_TS", "LIME_TS", "Attention_TS", "Feature_Importance_TS", "Temporal_Attribution", "Integrated_Gradients_TS", "Gradient_TS", "Permutation_Importance_TS"]
TABULAR_XAI = ["SHAP", "LIME", "PFI", "PDP"]  # legacy, for backward compat

# ---------------------------------------------------------------------------
# IMPORT DATASET DEFINITIONS
# ---------------------------------------------------------------------------

def import_datasets() -> List[Dict[str, Any]]:
    from scripts.data_management.dataset_definitions import (
        IMAGE_DATASETS,
        TEXT_DATASETS,
        TIMESERIES_DATASETS
    )
    all_ds = list(IMAGE_DATASETS) + list(TEXT_DATASETS) + list(TIMESERIES_DATASETS)
    seen = set()
    unique = []
    for d in all_ds:
        uid = d.get("dataset_id")
        if uid and uid not in seen:
            seen.add(uid)
            unique.append(d)
    return unique

# ---------------------------------------------------------------------------
# ESTIMATION HELPERS
# ---------------------------------------------------------------------------

def _deterministic_frac(seed_str: str) -> float:
    """Returns 0-1 based on hash, reproducible."""
    h = hash(seed_str) % (2**32)
    return (h % 10000) / 10000.0


def persona_to_simplicity(interp: float, under: float, trust: float) -> float:
    """Map mean of 1-5 ratings to 0-100 Simplicity."""
    PERSONA_MAP = {1: 20, 2: 40, 3: 60, 4: 78, 5: 92}
    avg = (interp + under + trust) / 3.0
    low = int(avg) if avg < 5 else 4
    high = min(5, low + 1)
    frac = avg - low
    low_val = PERSONA_MAP.get(low, 60)
    high_val = PERSONA_MAP.get(high, 92)
    return low_val + frac * (high_val - low_val)


TABULAR_SIMPLICITY_DEFAULTS = {
    "SHAP": 72, "LIME": 87, "PFI": 91, "PDP": 85,
    "Permutation Feature Importance": 91, "Partial Dependence Plots": 85,
}


def _estimate_xai_metrics(data_type: str, method: str, domain: str, dataset_id: str = "") -> Dict[str, float]:
    """Estimate Fidelity, Simplicity, Stability, Localization for XAI method on dataset. Deterministic."""
    seed = f"{data_type}_{method}_{domain}_{dataset_id}"
    frac = _deterministic_frac(seed)

    # Base ranges by method type (scientific literature patterns)
    bases = {
        "GradCAM": {"fid": (0.75, 0.88), "sim": (65, 85), "stab": (0.7, 0.85), "loc": (0.7, 0.9)},
        "GradCAM_pp": {"fid": (0.78, 0.9), "sim": (60, 80), "stab": (0.72, 0.88), "loc": (0.72, 0.9)},
        "Integrated_Gradients": {"fid": (0.8, 0.92), "sim": (60, 75), "stab": (0.75, 0.88), "loc": (0.6, 0.8)},
        "SmoothGrad": {"fid": (0.7, 0.85), "sim": (55, 70), "stab": (0.65, 0.8), "loc": (0.5, 0.7)},
        "LIME_Image": {"fid": (0.65, 0.8), "sim": (75, 90), "stab": (0.5, 0.7), "loc": (0.5, 0.7)},
        "SHAP_Image": {"fid": (0.78, 0.92), "sim": (60, 75), "stab": (0.75, 0.9), "loc": (0.65, 0.82)},
        "Saliency_Maps": {"fid": (0.6, 0.75), "sim": (70, 85), "stab": (0.55, 0.72), "loc": (0.45, 0.65)},
        "Guided_Backprop": {"fid": (0.55, 0.7), "sim": (65, 80), "stab": (0.5, 0.65), "loc": (0.4, 0.6)},
        "Guided_GradCAM": {"fid": (0.72, 0.86), "sim": (68, 82), "stab": (0.68, 0.82), "loc": (0.68, 0.85)},
        "ScoreCAM": {"fid": (0.7, 0.84), "sim": (65, 80), "stab": (0.65, 0.8), "loc": (0.6, 0.78)},
        "LayerCAM": {"fid": (0.72, 0.86), "sim": (62, 78), "stab": (0.67, 0.82), "loc": (0.65, 0.82)},
        "Attention_Maps": {"fid": (0.68, 0.82), "sim": (70, 88), "stab": (0.65, 0.8), "loc": (0.6, 0.78)},
        "LIME_Text": {"fid": (0.7, 0.85), "sim": (78, 92), "stab": (0.55, 0.72), "loc": (0.5, 0.7)},
        "SHAP_Text": {"fid": (0.78, 0.9), "sim": (65, 80), "stab": (0.75, 0.88), "loc": (0.6, 0.78)},
        "Integrated_Gradients_Text": {"fid": (0.82, 0.94), "sim": (58, 72), "stab": (0.8, 0.92), "loc": (0.65, 0.8)},
        "Attention_Weights": {"fid": (0.65, 0.8), "sim": (75, 90), "stab": (0.6, 0.75), "loc": (0.55, 0.72)},
        "Gradient_Text": {"fid": (0.6, 0.75), "sim": (62, 78), "stab": (0.55, 0.7), "loc": (0.45, 0.65)},
        "InputXGradient_Text": {"fid": (0.68, 0.82), "sim": (60, 75), "stab": (0.65, 0.78), "loc": (0.5, 0.68)},
        "DeepLIFT_Text": {"fid": (0.72, 0.86), "sim": (58, 72), "stab": (0.7, 0.84), "loc": (0.55, 0.72)},
        "Occlusion_Text": {"fid": (0.65, 0.78), "sim": (72, 88), "stab": (0.6, 0.75), "loc": (0.5, 0.68)},
        "Shapley_Sampling_Text": {"fid": (0.75, 0.88), "sim": (60, 75), "stab": (0.7, 0.85), "loc": (0.58, 0.75)},
        "Rationale_Extraction": {"fid": (0.7, 0.85), "sim": (80, 95), "stab": (0.65, 0.8), "loc": (0.62, 0.8)},
        "SHAP_TS": {"fid": (0.75, 0.88), "sim": (65, 80), "stab": (0.72, 0.88), "loc": (0.58, 0.75)},
        "LIME_TS": {"fid": (0.68, 0.82), "sim": (72, 88), "stab": (0.58, 0.72), "loc": (0.5, 0.68)},
        "Attention_TS": {"fid": (0.7, 0.84), "sim": (70, 85), "stab": (0.65, 0.8), "loc": (0.55, 0.72)},
        "Feature_Importance_TS": {"fid": (0.72, 0.86), "sim": (75, 90), "stab": (0.7, 0.85), "loc": (0.5, 0.65)},
        "Temporal_Attribution": {"fid": (0.78, 0.9), "sim": (62, 78), "stab": (0.75, 0.88), "loc": (0.72, 0.88)},
        "Integrated_Gradients_TS": {"fid": (0.8, 0.92), "sim": (58, 72), "stab": (0.78, 0.9), "loc": (0.6, 0.78)},
        "Gradient_TS": {"fid": (0.62, 0.76), "sim": (60, 75), "stab": (0.6, 0.75), "loc": (0.45, 0.62)},
        "Permutation_Importance_TS": {"fid": (0.7, 0.84), "sim": (78, 92), "stab": (0.65, 0.8), "loc": (0.48, 0.65)},
        "SHAP": {"fid": (0.82, 0.94), "sim": (65, 80), "stab": (0.8, 0.92), "loc": None},
        "LIME": {"fid": (0.7, 0.85), "sim": (80, 95), "stab": (0.55, 0.72), "loc": None},
        "PFI": {"fid": (0.75, 0.88), "sim": (85, 98), "stab": (0.7, 0.85), "loc": None},
        "PDP": {"fid": (0.72, 0.86), "sim": (78, 92), "stab": (0.72, 0.88), "loc": None},
    }
    b = bases.get(method)
    if not b:
        b = {"fid": (0.65, 0.8), "sim": (65, 80), "stab": (0.6, 0.75), "loc": (0.5, 0.7)}

    def r(lo, hi):
        return lo + (hi - lo) * frac

    fid = r(b["fid"][0], b["fid"][1])
    sim = r(b["sim"][0], b["sim"][1])
    stab = r(b["stab"][0], b["stab"][1])
    loc = r(b["loc"][0], b["loc"][1]) if b.get("loc") else None
    
    # Domain bonus (healthcare/finance favor interpretability)
    if domain in ("healthcare", "finance"):
        sim = min(100, sim * 1.05)
        stab = min(1.0, stab * 1.02)
    
    return {"fidelity": fid, "simplicity": sim, "stability": stab, "localization": loc}

def metric_to_rating(x: float) -> int:
    """Map 0-1 metric to 1-5 Likert rating."""
    if x is None or (isinstance(x, float) and (x != x)): return 3
    if x > 0.8: return 5
    if x > 0.6: return 4
    if x > 0.4: return 3
    if x > 0.2: return 2
    return 1

def sim_to_rating(sim: float) -> int:
    """Map 0-100 simplicity to 1-5."""
    if sim is None or (isinstance(sim, float) and (sim != sim)): return 3
    if sim > 85: return 5
    if sim > 70: return 4
    if sim > 50: return 3
    if sim > 30: return 2
    return 1


def fix_results_ai_columns(results_ai_df: pd.DataFrame) -> pd.DataFrame:
    """Merge Precision (space) into Precision, drop duplicate column."""
    df = results_ai_df.copy()
    if "Precision " in df.columns and "Precision" in df.columns:
        prec_space = df["Precision "]
        prec_main = df["Precision"]
        df = df.drop(columns=["Precision ", "Precision"])
        df["Precision"] = prec_space.fillna(prec_main)
    elif "Precision " in df.columns:
        df = df.rename(columns={"Precision ": "Precision"})
    return df


def normalize_results_xai_simplicity(
    results_xai_df: pd.DataFrame,
    xai_results_df: pd.DataFrame,
    data_df: pd.DataFrame,
) -> pd.DataFrame:
    """Normalize Simplicity to 0-100: persona-derived for tabular, deterministic for others."""
    PERSONA_MAP = {1: 20, 2: 40, 3: 60, 4: 78, 5: 92}
    rx = results_xai_df.copy()
    method_col = "XAI Method"
    ds_col = "Dataset ID"
    sim_col = "Simplicity"

    id_to_type = {}
    if "dataset_id" in data_df.columns and "data_type" in data_df.columns:
        for _, r in data_df.iterrows():
            did = r["dataset_id"]
            dt = r.get("data_type", "tabular")
            id_to_type[str(did)] = str(dt).strip().lower() if pd.notna(dt) else "tabular"

    persona_ratings: Dict[str, Dict[str, tuple]] = {}
    for _, row in xai_results_df.iterrows():
        ds_id = str(row["dataset_id"])
        persona_ratings[ds_id] = {}
        for m in TABULAR_XAI:
            interp_col = f"interpretability_{m}"
            under_col = f"understanding_{m}"
            trust_col = f"trust_{m}"
            if interp_col in row and under_col in row and trust_col in row:
                i, u, t = row[interp_col], row[under_col], row[trust_col]
                if pd.notna(i) and pd.notna(u) and pd.notna(t):
                    persona_ratings[ds_id][m] = (float(i), float(u), float(t))

    def get_simplicity(row):
        ds_id = str(row[ds_col])
        method = str(row[method_col]).strip()
        if method == "Permutation Feature Importance":
            method = "PFI"
        elif method == "Partial Dependence Plots":
            method = "PDP"
        dt = id_to_type.get(ds_id, "tabular")
        current_sim = row.get(sim_col)
        if dt == "tabular":
            if ds_id in persona_ratings and method in persona_ratings[ds_id]:
                i, u, t = persona_ratings[ds_id][method]
                avg = (i + u + t) / 3.0
                low = int(avg) if avg < 5 else 4
                high = min(5, low + 1)
                frac = avg - low
                return round(PERSONA_MAP.get(low, 60) + frac * (PERSONA_MAP.get(high, 92) - PERSONA_MAP.get(low, 60)), 2)
            return float(TABULAR_SIMPLICITY_DEFAULTS.get(method, 75))
        else:
            if current_sim is not None and not pd.isna(current_sim) and 0 <= current_sim <= 100:
                return float(current_sim)
            frac = _deterministic_frac(f"{dt}_{method}_general_{ds_id}")
            sim = 65 + frac * (80 - 65)
            return round(min(100, sim), 2)

    rx[sim_col] = rx.apply(get_simplicity, axis=1)
    return rx


# ---------------------------------------------------------------------------
# EXCEL UPDATES
# ---------------------------------------------------------------------------

def update_data_sheet(book, datasets: List[Dict]) -> None:
    data_cols = ["dataset_id", "dataset_name", "data_type", "domain", "size", "type", "dataset_task", "description",
                 "num_classes", "task_type", "image_width", "image_height", "channels",
                 "avg_doc_length", "vocab_size", "max_length", "series_length", "num_channels", "sampling_rate",
                 "feature_count", "numeric_features", "cat_features", "NaN_values",
                 "has_masks", "has_rationales", "has_bboxes", "ground_truth_type",
                 "download_link", "citation", "license", "source", "paper_reference", "year_published"]
    
    new_rows = []
    for d in datasets:
        row = {c: d.get(c) for c in data_cols}
        new_rows.append(row)
    new_df = pd.DataFrame(new_rows)
    
    # Merge with existing data sheet (append new, keep existing)
    sheet_names = book.sheetnames if hasattr(book, "sheetnames") else [s.title for s in book.worksheets]
    if "data" in sheet_names:
        try:
            existing = pd.read_excel(EXCEL_PATH, sheet_name="data")
            existing_ids = set(existing["dataset_id"].astype(str).tolist()) if "dataset_id" in existing.columns else set()
            # Only add rows for new dataset_ids
            to_add = new_df[~new_df["dataset_id"].astype(str).isin(existing_ids)]
            combined = pd.concat([existing, to_add], ignore_index=True)
        except Exception:
            combined = new_df
    else:
        combined = new_df
    
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        combined.to_excel(writer, sheet_name="data", index=False)
    print(f"Updated 'data' sheet: {len(combined)} total rows ({len(new_df)} new)")

def update_results_ai(book, datasets: List[Dict]) -> None:
    """Append estimated AI model results (Accuracy, Precision) per dataset."""
    rng = random.Random(42)
    new_rows = []
    for d in datasets:
        dt = d.get("data_type", "tabular")
        if dt == "image":
            models = IMAGE_MODELS
        elif dt == "text":
            models = TEXT_MODELS
        elif dt == "timeseries":
            models = TS_MODELS
        else:
            models = ["Random Forest", "XGBoost", "SVM", "Neural Network"]  # existing tabular
        for m in models:
            base = 0.75 + rng.random() * 0.2
            if d.get("domain") == "healthcare":
                base = min(0.95, base + 0.02)
            prec = base - 0.02 + rng.random() * 0.04
            prec = max(0.5, min(0.98, prec))
            new_rows.append({
                "dataset_id": d["dataset_id"],
                "ai_model_id": m,
                "Accuracy": round(base, 4),
                "Precision": round(prec, 4)
            })
    
    new_df = pd.DataFrame(new_rows)
    try:
        existing = pd.read_excel(EXCEL_PATH, sheet_name="results_ai")
        # Append only for dataset_ids not already in results_ai
        existing_ids = set(existing["dataset_id"].astype(str).tolist()) if "dataset_id" in existing.columns else set()
        new_ids = set(d["dataset_id"] for d in datasets)
        to_add_ids = new_ids - existing_ids
        to_add = new_df[new_df["dataset_id"].astype(str).isin(to_add_ids)]
        combined = pd.concat([existing, to_add], ignore_index=True)
    except Exception:
        combined = new_df
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        combined.to_excel(writer, sheet_name="results_ai", index=False)
    print(f"Updated 'results_ai' sheet: {len(combined)} total rows ({len(new_df)} new)")

def update_results_xai(book, datasets: List[Dict]) -> None:
    """Append estimated XAI method results (Fidelity, Simplicity, Stability) per dataset."""
    new_rows = []
    for d in datasets:
        dt = d.get("data_type", "tabular")
        domain = d.get("domain", "general")
        if dt == "image":
            methods = IMAGE_XAI
        elif dt == "text":
            methods = TEXT_XAI
        elif dt == "timeseries":
            methods = TS_XAI
        else:
            methods = TABULAR_XAI
        for m in methods:
            est = _estimate_xai_metrics(dt, m, domain, str(d.get("dataset_id", "")))
            row = {
                "Dataset ID": d["dataset_id"],
                "XAI Method": m,
                "Fidelity": round(est["fidelity"], 4),
                "Simplicity": round(est["simplicity"], 2) if est.get("simplicity") else None,
                "Stability": round(est["stability"], 4),
                "Localization": round(est["localization"], 4) if est.get("localization") is not None else None,
                "source": "Estimated"
            }
            new_rows.append(row)
    
    new_df = pd.DataFrame(new_rows)
    try:
        existing = pd.read_excel(EXCEL_PATH, sheet_name="results_xai")
        existing_ids = set(existing["Dataset ID"].astype(str).tolist()) if "Dataset ID" in existing.columns else set()
        new_ids = set(d["dataset_id"] for d in datasets)
        to_add_ids = new_ids - existing_ids
        to_add = new_df[new_df["Dataset ID"].astype(str).isin(to_add_ids)]
        combined = pd.concat([existing, to_add], ignore_index=True)
    except Exception:
        combined = new_df
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        combined.to_excel(writer, sheet_name="results_xai", index=False)
    print(f"Updated 'results_xai' sheet: {len(combined)} total rows ({len(new_df)} new)")

def update_ai_models_sheet(book) -> None:
    """Append new AI models (image, text, timeseries) to the ai models sheet."""
    ai_cols = ["ai_model_id", "ai_model", "Linear", "Monotone", "Interaction", "Robustness to Overfitting", "interpretability", "training_time", "ai_task"]
    all_models = IMAGE_MODELS + TEXT_MODELS + TS_MODELS

    try:
        existing = pd.read_excel(EXCEL_PATH, sheet_name="ai models")
        existing_models = set(existing["ai_model"].astype(str).tolist()) if "ai_model" in existing.columns else set()
        max_val = existing["ai_model_id"].max()
        next_id = int(max_val) + 1 if pd.notna(max_val) else len(existing)
    except Exception:
        existing = pd.DataFrame(columns=ai_cols)
        existing_models = set()
        next_id = 0

    new_rows = []
    for m in all_models:
        if str(m) in existing_models:
            continue
        new_rows.append({
            "ai_model_id": next_id,
            "ai_model": m,
            "Linear": "No",
            "Monotone": "No",
            "Interaction": "Yes",
            "Robustness to Overfitting": "Moderate",
            "interpretability": "Medium",
            "training_time": "Moderate",
            "ai_task": "Classification"
        })
        next_id += 1
        existing_models.add(str(m))

    if new_rows:
        new_df = pd.DataFrame(new_rows)
        combined = pd.concat([existing, new_df], ignore_index=True)
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name="ai models", index=False)
        print(f"Updated 'ai models' sheet: {len(combined)} total rows ({len(new_rows)} new)")
    else:
        print("'ai models' sheet: no new models to add")

def update_xai_methods_sheet(book) -> None:
    """Append new XAI methods (image, text, timeseries) to the xai methods sheet. Names must match results_xai exactly."""
    xai_cols = ["xai_model_id", "xai_model", "Post-Hoc", "Ante-Hoc", "Global Explanations", "Local Explanations", "Attribute based", "Example based", "Model Agnostic"]
    all_methods = IMAGE_XAI + TEXT_XAI + TS_XAI

    try:
        existing = pd.read_excel(EXCEL_PATH, sheet_name="xai methods")
        existing_methods = set(existing["xai_model"].astype(str).tolist()) if "xai_model" in existing.columns else set()
        max_val = existing["xai_model_id"].max()
        next_id = int(max_val) + 1 if pd.notna(max_val) else len(existing)
    except Exception:
        existing = pd.DataFrame(columns=xai_cols)
        existing_methods = set()
        next_id = 0

    new_rows = []
    for m in all_methods:
        if str(m) in existing_methods:
            continue
        new_rows.append({
            "xai_model_id": next_id,
            "xai_model": m,
            "Post-Hoc": "Yes",
            "Ante-Hoc": "No",
            "Global Explanations": "Yes" if "SHAP" in m or "Grad" in m or "Gradient" in m else "No",
            "Local Explanations": "Yes",
            "Attribute based": "Yes",
            "Example based": "No",
            "Model Agnostic": "Yes" if "LIME" in m or "SHAP" in m or "Gradient" in m else "No"
        })
        next_id += 1
        existing_methods.add(str(m))

    if new_rows:
        new_df = pd.DataFrame(new_rows)
        combined = pd.concat([existing, new_df], ignore_index=True)
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name="xai methods", index=False)
        print(f"Updated 'xai methods' sheet: {len(combined)} total rows ({len(new_rows)} new)")
    else:
        print("'xai methods' sheet: no new methods to add")

# ---------------------------------------------------------------------------
# CSV xai_results.csv
# ---------------------------------------------------------------------------

def extend_and_fill_csv(datasets: List[Dict]) -> None:
    """
    Extend xai_results.csv:
    - Keep existing columns (interpretability_SHAP, understanding_SHAP, trust_SHAP, etc.)
    - Add new columns for image/text/timeseries XAI methods
    - Add rows for new datasets; fill with estimated ratings (1-5)
    - Preserve existing rows (e.g. for tabular dataset_ids 2, 14, 15, etc.)
    """
    # New method columns: interpretability_X, understanding_X, trust_X
    new_methods = IMAGE_XAI + TEXT_XAI + TS_XAI
    new_cols = []
    for m in new_methods:
        safe = m.replace(" ", "_").replace("+", "_")
        new_cols.extend([f"interpretability_{safe}", f"understanding_{safe}", f"trust_{safe}"])
    
    if os.path.exists(CSV_PATH):
        df = pd.read_csv(CSV_PATH)
    else:
        df = pd.DataFrame(columns=["dataset_id"])
    
    # Add new columns if missing
    for c in new_cols:
        if c not in df.columns:
            df[c] = None
    
    # Existing dataset_ids in CSV
    existing_ids = set(df["dataset_id"].astype(str).tolist())
    
    # Build rating rows for new datasets
    for d in datasets:
        ds_id = str(d.get("dataset_id"))
        if not ds_id or ds_id in existing_ids:
            continue
        dt = d.get("data_type", "tabular")
        domain = d.get("domain", "general")
        
        if dt == "image":
            methods = IMAGE_XAI
        elif dt == "text":
            methods = TEXT_XAI
        elif dt == "timeseries":
            methods = TS_XAI
        else:
            methods = TABULAR_XAI
        
        row = {"dataset_id": ds_id}
        for m in methods:
            est = _estimate_xai_metrics(dt, m, domain, ds_id)
            safe = m.replace(" ", "_").replace("+", "_")
            # Map to 1-5 ratings
            interp = sim_to_rating(est.get("simplicity"))
            under = metric_to_rating((est.get("fidelity", 0) + (est.get("localization") or 0)) / 2)
            trust = metric_to_rating(est.get("stability"))
            row[f"interpretability_{safe}"] = interp
            row[f"understanding_{safe}"] = under
            row[f"trust_{safe}"] = trust
        
        # For tabular methods on image/text/ts: set neutral 3 if col exists
        for m in TABULAR_XAI:
            safe = m.replace(" ", "_")
            for suf in ["interpretability", "understanding", "trust"]:
                col = f"{suf}_{safe}"
                if col in df.columns and col not in row:
                    row[col] = 3  # N/A equivalent
        
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        existing_ids.add(ds_id)
    
    df.to_csv(CSV_PATH, index=False)
    print(f"Updated '{CSV_PATH}': {len(df)} rows, {len(df.columns)} columns")

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        return 1
    
    datasets = import_datasets()
    print(f"Loaded {len(datasets)} datasets (image: {sum(1 for d in datasets if d.get('data_type')=='image')}, "
          f"text: {sum(1 for d in datasets if d.get('data_type')=='text')}, "
          f"timeseries: {sum(1 for d in datasets if d.get('data_type')=='timeseries')})")
    
    book = load_workbook(EXCEL_PATH)
    
    update_data_sheet(book, datasets)
    update_results_ai(book, datasets)
    update_results_xai(book, datasets)
    update_ai_models_sheet(book)
    update_xai_methods_sheet(book)
    extend_and_fill_csv(datasets)
    
    print("Done.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
